import csv
from operator import itemgetter
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter


class report:
    def __init__(self, vacancies_count_by_years, vacancies_count_by_years_for_profession, salary_by_years,
                 salary_by_years_for_profession, vacancies_count_by_cities, vacancies_share_by_cities, salary_by_cities,
                 profession):
        self.vacancies_count_by_years = vacancies_count_by_years
        self.vacancies_count_by_years_for_profession = vacancies_count_by_years_for_profession
        self.salary_by_years = salary_by_years
        self.salary_by_years_for_profession = salary_by_years_for_profession
        self.vacancies_count_by_cities = vacancies_count_by_cities
        self.vacancies_share_by_cities = vacancies_share_by_cities
        self.salary_by_cities = salary_by_cities
        self.side = Side(style='thin', color="000000")
        self.thin_border = Border(left=self.side, right=self.side, top=self.side, bottom=self.side)
        self.profession = profession

    def make_column_stylied(self, ws, index_column, max_row_index):
        max_value = -6
        for row in range(1, max_row_index):
            cell = ws.cell(column=index_column, row=row)
            if max_value < len(str(cell.value)):
                max_value = len(str(cell.value))
            cell.border = self.thin_border
        ws.column_dimensions[get_column_letter(index_column)].width = max_value + 2

    def fill_column_by_years(self, ws, column_index, dictionary_values, min_year, max_year):
        max_row_index = max_year - min_year
        for row, value in zip([index for index in range(2, max_row_index + 2)],
                              [dictionary_values[value] for value in range(min_year, max_year)]):
            ws.cell(row=row, column=column_index, value=value)

    def fill_column_by_cities(self, ws, column_index, max_row_index, dictionary_values, keys):
        for row, value in zip([index for index in range(2, max_row_index + 2)],
                              [dictionary_values[key] for key in keys]):
            ws.cell(row=row, column=column_index, value=value)

    def make_ws_by_years(self, ws):
        min_year, max_year = 2007, 2022
        max_column_index = 5
        max_row_index = max_year - min_year
        ws.title = "Статистика по годам"
        self.make_titles(max_column_index,
                         ["Год", "Средняя зарплата", f"Средняя зарплата - {self.profession}", "Количество вакансий",
                          f"Количество вакансий - {self.profession}"],
                         ws)
        self.fill_column_by_years(ws, 1,
                                  {x: y for x, y in zip(range(min_year, max_year + 1), range(min_year, max_year + 1))},
                                  min_year, max_year + 1)
        self.fill_column_by_years(ws, 2, self.salary_by_years, min_year, max_year + 1)
        self.fill_column_by_years(ws, 3, self.salary_by_years_for_profession, min_year, max_year + 1)
        self.fill_column_by_years(ws, 4, self.vacancies_count_by_years, min_year, max_year + 1)
        self.fill_column_by_years(ws, 5, self.vacancies_count_by_years_for_profession, min_year, max_year + 1)
        for i in range(1, 6):
            self.make_column_stylied(ws, i, max_row_index + 3)

    def make_ws_by_cities(self, ws):
        max_column_index = 5
        max_row_index = 11
        ws.title = "Статистика по городам"
        self.make_titles(max_column_index, ["Город", "Уроввень зарплат", "", "Город", "Доля вакансий"], ws)
        keys = self.salary_by_cities.keys()
        self.fill_column_by_cities(ws, 1, max_row_index, {x: y for x, y in zip(keys, keys)}, keys)
        self.fill_column_by_cities(ws, 2, max_row_index, self.salary_by_cities, keys)
        keys = self.vacancies_share_by_cities.keys()
        self.fill_column_by_cities(ws, 4, max_row_index, {x: y for x, y in zip(keys, keys)}, keys)
        self.fill_column_by_cities(ws, 5, max_row_index, self.vacancies_share_by_cities, keys)
        for i in range(1, 6):
            self.make_column_stylied(ws, i, max_row_index + 1)
        ws.column_dimensions[get_column_letter(3)].width = 2
        for cell in ws['E']:
            cell.number_format = '0.00%'

    def make_titles(self, max_column_index, names, ws):
        for cols_index, name in zip([i for i in range(1, max_column_index + 1)], names):
            ws.cell(row=1, column=cols_index, value=name).font = Font(bold=True)

    def generate_excel(self):
        wb = Workbook()
        ws_by_years = wb.active
        ws_by_cities = wb.create_sheet()
        self.make_ws_by_years(ws_by_years)
        self.make_ws_by_cities(ws_by_cities)
        wb.save("report.xlsx")


class Vacancy:
    def __init__(self, dictionary):
        self.currency_to_rub = {"AZN": 35.68,
                                "BYR": 23.91,
                                "EUR": 59.90,
                                "GEL": 21.74,
                                "KGS": 0.76,
                                "KZT": 0.13,
                                "RUR": 1,
                                "UAH": 1.64,
                                "USD": 60.66,
                                "UZS": 0.0055, }
        self.name = dictionary["name"]
        self.salary = (float(dictionary["salary_from"]) + float(dictionary["salary_to"])) / 2 * self.currency_to_rub[
            dictionary["salary_currency"]]
        self.area_name = dictionary["area_name"]
        self.published_at = int(dictionary["published_at"][:4])


class DataSet:
    def __init__(self, file_name, profession):
        self.file_name = file_name
        self.profession = profession
        csv_read = self.csv_reader()
        dictionaries = self.csv_filer(csv_read[1], csv_read[0])
        vacancies_list = []
        for dictionary in dictionaries:
            vacancies_list.append(Vacancy(dictionary))
        self.vacancies_objects = vacancies_list
        self.vacancies_count_by_years = self.count_vacancies_by_years()
        self.vacancies_count_by_years_for_profession = self.count_profession_vacancies_by_years()
        self.salary_by_years = self.get_salary_by_years()
        self.salary_by_years_for_profession = self.get_profession_salary_by_years()
        self.vacancies_count_by_cities = self.count_vacancies_by_cities()
        self.vacancies_share_by_cities = self.get_vacancies_share_by_cities()
        self.salary_by_cities = self.get_salary_by_cities()

    def check_rows_count(self, rows_count):
        if rows_count == 0:
            print("Пустой файл")
            exit()
        if rows_count == 1:
            print("Нет данных")
            exit()

    def take_ten_items(self, dictionary):
        result_dictionary = {}
        for key, i in zip(dictionary, [i for i in range(10)]):
            result_dictionary[key] = round(dictionary[key], 4)
        return result_dictionary

    def csv_reader(self):
        vacancies, headlines = [], []
        length, rows_count = 0, 0
        first = True
        with open(self.file_name, encoding="utf-8-sig") as File:
            file = csv.reader(File)
            for row in file:
                rows_count += 1
                if first:
                    length = len(row)
                    headlines = row
                    first = False
                else:
                    is_break = False
                    if length != len(row): is_break = True
                    for word in row:
                        if word == "": is_break = True
                    if is_break: continue
                    vacancies.append(row)
        self.check_rows_count(rows_count)
        return headlines, vacancies

    def csv_filer(self, reader, list_naming):
        dictionaries = []
        for vacancy in reader:
            dictionary = {}
            for name, item in zip(list_naming, vacancy):
                dictionary[name] = item
            dictionaries.append(dictionary)
        return dictionaries

    def get_vacancies_share_by_cities(self):
        dictionary = {}
        for key in self.vacancies_count_by_cities:
            if self.vacancies_count_by_cities[key] / len(self.vacancies_objects) >= 0.01:
                dictionary[key] = self.vacancies_count_by_cities[key] / len(self.vacancies_objects)
        return self.take_ten_items(dict(sorted(dictionary.items(), key=itemgetter(1), reverse=True)))

    def get_salary_by_cities(self):
        dictionary = {}
        for vacancy in self.vacancies_objects:
            if self.vacancies_count_by_cities[vacancy.area_name] / len(self.vacancies_objects) < 0.01:
                continue
            dictionary[vacancy.area_name] = (
                dictionary[
                    vacancy.area_name] + vacancy.salary if vacancy.area_name in dictionary else vacancy.salary)
        for key in dictionary:
            dictionary[key] = int(dictionary[key] / self.vacancies_count_by_cities[key])
        return self.take_ten_items(dict(sorted(dictionary.items(), key=itemgetter(1), reverse=True)))

    def count_vacancies_by_years(self):
        dictionary = {}
        for vacancy in self.vacancies_objects:
            dictionary[vacancy.published_at] = (
                dictionary[vacancy.published_at] + 1 if vacancy.published_at in dictionary else 1)
        dictionary = dict(sorted(dictionary.items(), key=itemgetter(0)))
        return dictionary

    def count_profession_vacancies_by_years(self):
        dictionary = {}
        for vacancy in self.vacancies_objects:
            if self.profession not in vacancy.name:
                continue
            dictionary[vacancy.published_at] = (
                dictionary[vacancy.published_at] + 1 if vacancy.published_at in dictionary else 1)
        dictionary = dict(sorted(dictionary.items(), key=itemgetter(0)))
        if len(dictionary) == 0:
            dictionary[2022] = 0
        return dictionary

    def get_salary_by_years(self):
        dictionary = {}
        for vacancy in self.vacancies_objects:
            dictionary[vacancy.published_at] = (
                dictionary[
                    vacancy.published_at] + vacancy.salary if vacancy.published_at in dictionary else vacancy.salary)
        for key in dictionary:
            dictionary[key] = int(dictionary[key] / self.vacancies_count_by_years[key])
        return dict(sorted(dictionary.items(), key=itemgetter(0)))

    def get_profession_salary_by_years(self):
        dictionary = {}
        for vacancy in self.vacancies_objects:
            if self.profession not in vacancy.name:
                continue
            dictionary[vacancy.published_at] = (
                dictionary[
                    vacancy.published_at] + vacancy.salary if vacancy.published_at in dictionary else vacancy.salary)
        for key in dictionary:
            dictionary[key] = int(dictionary[key] / self.vacancies_count_by_years_for_profession[key])
        dictionary = dict(sorted(dictionary.items(), key=itemgetter(0)))
        if len(dictionary) == 0:
            dictionary[2022] = 0
        return dictionary

    def count_vacancies_by_cities(self):
        dictionary = {}
        for vacancy in self.vacancies_objects:
            dictionary[vacancy.area_name] = (
                dictionary[
                    vacancy.area_name] + 1 if vacancy.area_name in dictionary else 1)
        return dictionary


file_name = input("Введите название файла: ")
profession = input("Введите название профессии: ")
dataset = DataSet(file_name, profession)
report(dataset.vacancies_count_by_years, dataset.vacancies_count_by_years_for_profession, dataset.salary_by_years,
       dataset.salary_by_years_for_profession,
       dataset.vacancies_count_by_cities, dataset.vacancies_share_by_cities, dataset.salary_by_cities,
       dataset.profession).generate_excel()
